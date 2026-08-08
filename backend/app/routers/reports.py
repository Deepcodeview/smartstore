"""
routers/reports.py — PDF & Excel report export for Retail AI
GET /reports/{job_id}/pdf   → downloadable PDF report
GET /reports/{job_id}/excel → downloadable Excel workbook
GET /reports/shift/pdf      → shift summary PDF (all completed jobs)
"""
import io
import datetime
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database.db import get_db
from app.database.models import AnalyticsJob, CrossingEvent, AlertLog, JobStatus

router = APIRouter(prefix="/reports", tags=["reports"])


def _get_job_or_404(job_id: str, db: Session) -> AnalyticsJob:
    job = db.query(AnalyticsJob).filter(AnalyticsJob.job_id == job_id).first()
    if not job or not job.result:
        raise HTTPException(404, "Job not found or not completed yet.")
    return job


# ── PDF Export ────────────────────────────────────────────────────────────────

@router.get("/{job_id}/pdf")
def export_pdf(job_id: str, db: Session = Depends(get_db)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    job = _get_job_or_404(job_id, db)
    r   = job.result

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"],
                                 fontSize=20, textColor=colors.HexColor("#0057ff"),
                                 spaceAfter=4, alignment=TA_CENTER)
    sub_style   = ParagraphStyle("sub", parent=styles["Normal"],
                                 fontSize=10, textColor=colors.HexColor("#94a3b8"),
                                 alignment=TA_CENTER, spaceAfter=16)
    h2_style    = ParagraphStyle("h2", parent=styles["Heading2"],
                                 fontSize=13, textColor=colors.HexColor("#0f1923"),
                                 spaceBefore=14, spaceAfter=6)
    body_style  = ParagraphStyle("body", parent=styles["Normal"],
                                 fontSize=10, textColor=colors.HexColor("#4a5568"),
                                 leading=16)

    def tbl(data, col_widths=None):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#0057ff")),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0), 10),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#f8fafc"), colors.white]),
            ("FONTSIZE",    (0, 1), (-1, -1), 9),
            ("TEXTCOLOR",   (0, 1), (-1, -1), colors.HexColor("#4a5568")),
            ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#e8ecf0")),
            ("ROWPADDING",  (0, 0), (-1, -1), 6),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return t

    story = []

    # Header
    story.append(Paragraph("🏪 RetailVision AI", title_style))
    story.append(Paragraph(f"Store Analytics Report — {job.filename}", sub_style))
    story.append(Paragraph(
        f"Generated: {datetime.datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')}  |  Job ID: {job_id[:8]}…",
        sub_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e8ecf0")))
    story.append(Spacer(1, 12))

    # Footfall Summary
    story.append(Paragraph("1. Footfall Summary", h2_style))
    footfall_data = [
        ["Metric", "Value"],
        ["Total Unique People",  str(r.get("total_unique_people", 0))],
        ["Total Entries",        str(r.get("entries", 0))],
        ["Total Exits",          str(r.get("exits", 0))],
        ["Currently Inside",     str(r.get("currently_inside", 0))],
        ["Peak Crowd Count",     str(r.get("peak_crowd", {}).get("count", 0))],
        ["Peak Crowd Time",      f"{r.get('peak_crowd', {}).get('time_sec', 0):.1f}s"],
        ["Avg People / Frame",   str(r.get("avg_people_per_frame", 0))],
    ]
    story.append(tbl(footfall_data, [8*cm, 8*cm]))
    story.append(Spacer(1, 10))

    # Dwell Time
    dwell = r.get("dwell", {})
    if dwell:
        story.append(Paragraph("2. Dwell Time Analysis", h2_style))
        dwell_data = [["Metric", "Value"]]
        for k, v in dwell.items():
            dwell_data.append([k.replace("_", " ").title(), str(round(v, 2)) if isinstance(v, float) else str(v)])
        story.append(tbl(dwell_data, [8*cm, 8*cm]))
        story.append(Spacer(1, 10))

    # Zone Analytics
    zones = r.get("zones", {})
    if zones:
        story.append(Paragraph("3. Zone Analytics", h2_style))
        unique_v = zones.get("unique_visitors", {})
        avg_pf   = zones.get("avg_per_frame", {})
        zone_data = [["Zone", "Unique Visitors", "Avg / Frame"]]
        for z in unique_v:
            zone_data.append([z, str(unique_v.get(z, 0)), str(round(avg_pf.get(z, 0), 2))])
        story.append(tbl(zone_data, [6*cm, 5*cm, 5*cm]))
        story.append(Paragraph(
            f"Most Popular Zone: <b>{zones.get('most_popular', 'N/A')}</b>",
            body_style))
        story.append(Spacer(1, 10))

    # Shelf Status
    story.append(Paragraph("4. Shelf Stock Status", h2_style))
    shelf_data = [
        ["Status", "Value"],
        ["Final Shelf Status", r.get("shelf_status", "N/A")],
    ]
    breakdown = r.get("shelf_breakdown", {})
    for k, v in breakdown.items():
        shelf_data.append([f"  {k}", str(v) + " frames"])
    story.append(tbl(shelf_data, [8*cm, 8*cm]))
    story.append(Spacer(1, 10))

    # Alerts
    alerts = db.query(AlertLog).filter(AlertLog.job_id == job_id).all()
    if alerts:
        story.append(Paragraph("5. Alerts Log", h2_style))
        alert_data = [["Severity", "Message", "Time (s)"]]
        for a in alerts[:30]:
            alert_data.append([a.severity, a.message[:60], str(round(a.timestamp_sec, 1))])
        story.append(tbl(alert_data, [3*cm, 10*cm, 3*cm]))
        story.append(Spacer(1, 10))

    # Video Meta
    meta = r.get("video_meta", {})
    if meta:
        story.append(Paragraph("6. Video Metadata", h2_style))
        meta_data = [
            ["Property", "Value"],
            ["Resolution",  f"{meta.get('width', 0)}×{meta.get('height', 0)}"],
            ["FPS",         str(meta.get("fps", 0))],
            ["Total Frames", str(meta.get("total_frames", 0))],
            ["Processing Time", f"{r.get('processing_time_sec', 0)}s"],
        ]
        story.append(tbl(meta_data, [8*cm, 8*cm]))

    doc.build(story)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=retail_report_{job_id[:8]}.pdf"},
    )


# ── Excel Export ──────────────────────────────────────────────────────────────

@router.get("/{job_id}/excel")
def export_excel(job_id: str, db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    job = _get_job_or_404(job_id, db)
    r   = job.result

    wb = Workbook()

    BRAND   = "0057FF"
    HEADER  = "F8FAFC"
    BORDER  = "E8ECF0"

    def _hdr_font():  return Font(bold=True, color="FFFFFF", size=10)
    def _hdr_fill():  return PatternFill("solid", fgColor=BRAND)
    def _row_fill(i): return PatternFill("solid", fgColor="F8FAFC" if i % 2 == 0 else "FFFFFF")
    def _thin_border():
        s = Side(style="thin", color=BORDER)
        return Border(left=s, right=s, top=s, bottom=s)
    def _center(): return Alignment(horizontal="center", vertical="center")

    def write_sheet(ws, title, headers, rows):
        ws.title = title
        ws.append(headers)
        for cell in ws[1]:
            cell.font  = _hdr_font()
            cell.fill  = _hdr_fill()
            cell.alignment = _center()
            cell.border = _thin_border()
        for i, row in enumerate(rows, start=2):
            ws.append(row)
            for cell in ws[i]:
                cell.fill   = _row_fill(i)
                cell.border = _thin_border()
                cell.alignment = Alignment(vertical="center")
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # Sheet 1: Summary
    ws1 = wb.active
    write_sheet(ws1, "Summary", ["Metric", "Value"], [
        ["File",              job.filename],
        ["Job ID",            job_id],
        ["Status",            job.status],
        ["Created At",        str(job.created_at)],
        ["Completed At",      str(job.completed_at)],
        ["Total Unique People", r.get("total_unique_people", 0)],
        ["Entries",           r.get("entries", 0)],
        ["Exits",             r.get("exits", 0)],
        ["Currently Inside",  r.get("currently_inside", 0)],
        ["Peak Crowd",        r.get("peak_crowd", {}).get("count", 0)],
        ["Peak Time (s)",     r.get("peak_crowd", {}).get("time_sec", 0)],
        ["Avg People/Frame",  r.get("avg_people_per_frame", 0)],
        ["Shelf Status",      r.get("shelf_status", "N/A")],
        ["Processing Time (s)", r.get("processing_time_sec", 0)],
    ])

    # Sheet 2: Zone Analytics
    ws2 = wb.create_sheet("Zone Analytics")
    zones = r.get("zones", {})
    unique_v = zones.get("unique_visitors", {})
    avg_pf   = zones.get("avg_per_frame", {})
    zone_rows = [[z, unique_v.get(z, 0), round(avg_pf.get(z, 0), 2)] for z in unique_v]
    write_sheet(ws2, "Zone Analytics", ["Zone", "Unique Visitors", "Avg People/Frame"], zone_rows)

    # Sheet 3: Dwell Time
    ws3 = wb.create_sheet("Dwell Time")
    dwell = r.get("dwell", {})
    dwell_rows = [[k.replace("_", " ").title(), round(v, 3) if isinstance(v, float) else v]
                  for k, v in dwell.items()]
    write_sheet(ws3, "Dwell Time", ["Metric", "Value"], dwell_rows)

    # Sheet 4: Crossing Events
    ws4 = wb.create_sheet("Crossing Events")
    events = db.query(CrossingEvent).filter(CrossingEvent.job_id == job_id).all()
    crossing_rows = [[e.global_id, e.event_type, round(e.timestamp_sec, 2), str(e.wall_time)]
                     for e in events]
    write_sheet(ws4, "Crossing Events", ["Person ID", "Event Type", "Time (s)", "Wall Time"], crossing_rows)

    # Sheet 5: Alerts
    ws5 = wb.create_sheet("Alerts")
    alerts = db.query(AlertLog).filter(AlertLog.job_id == job_id).all()
    alert_rows = [[a.severity, a.message, round(a.timestamp_sec, 2), str(a.wall_time)]
                  for a in alerts]
    write_sheet(ws5, "Alerts", ["Severity", "Message", "Time (s)", "Wall Time"], alert_rows)

    # Sheet 6: Crowd Timeline
    ws6 = wb.create_sheet("Crowd Timeline")
    timeline = r.get("timeline", [])
    tl_rows = [[t["frame"], t["time_sec"], t["count"]] for t in timeline]
    write_sheet(ws6, "Crowd Timeline", ["Frame", "Time (s)", "People Count"], tl_rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=retail_report_{job_id[:8]}.xlsx"},
    )


# ── Shift Summary PDF (all completed jobs) ────────────────────────────────────

@router.get("/shift/pdf")
def export_shift_pdf(db: Session = Depends(get_db)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.enums import TA_CENTER

    jobs = db.query(AnalyticsJob).filter(AnalyticsJob.status == JobStatus.COMPLETED).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("t", parent=styles["Title"], fontSize=18,
                             textColor=colors.HexColor("#0057ff"), alignment=TA_CENTER)
    sub_s   = ParagraphStyle("s", parent=styles["Normal"], fontSize=10,
                             textColor=colors.HexColor("#94a3b8"), alignment=TA_CENTER)
    h2_s    = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12,
                             textColor=colors.HexColor("#0f1923"), spaceBefore=12, spaceAfter=6)

    story = [
        Paragraph("🏪 RetailVision AI — Shift Summary", title_s),
        Paragraph(f"All Completed Jobs  |  {datetime.date.today().strftime('%d %b %Y')}", sub_s),
        HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e8ecf0")),
        Spacer(1, 12),
    ]

    total_entries = total_exits = total_people = 0
    rows = [["File", "Entries", "Exits", "Peak", "Shelf", "Completed"]]
    for j in jobs:
        r = j.result or {}
        e  = r.get("entries", 0)
        ex = r.get("exits", 0)
        p  = r.get("peak_crowd", {}).get("count", 0) if isinstance(r.get("peak_crowd"), dict) else 0
        total_entries += e
        total_exits   += ex
        total_people  += r.get("total_unique_people", 0)
        rows.append([
            j.filename[:30],
            str(e), str(ex), str(p),
            r.get("shelf_status", "N/A"),
            str(j.completed_at)[:16] if j.completed_at else "—",
        ])

    t = Table(rows, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#0057ff")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f8fafc"), colors.white]),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#e8ecf0")),
        ("ROWPADDING",  (0, 0), (-1, -1), 5),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))

    story.append(Paragraph("Totals", h2_s))
    totals = Table([
        ["Total Entries", "Total Exits", "Total Unique People", "Jobs Completed"],
        [str(total_entries), str(total_exits), str(total_people), str(len(jobs))],
    ])
    totals.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#f0fdf4")),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 10),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#bbf7d0")),
        ("ROWPADDING",  (0, 0), (-1, -1), 8),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(totals)

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=shift_summary_{datetime.date.today()}.pdf"},
    )
